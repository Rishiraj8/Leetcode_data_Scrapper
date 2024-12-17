import requests
import json
import time
from openpyxl import load_workbook
from openpyxl import Workbook
from google.colab import files


url = "https://leetcode.com/graphql"

uploaded = files.upload()

wb = load_workbook(filename=list(uploaded.keys())[0])
sheet = wb.active

new_wb = Workbook()
new_sheet = new_wb.active

columns = ['S no', 'Roll no', 'Name', 'Class', 'LeetCode ID', 'LeetCode Profile Link', 'Easy Solved', 'Medium Solved', 'Hard Solved', 'Attended Contests Count', 'Rating']
new_sheet.append(columns)
ROWS = sheet.max_row

for i in range(2, ROWS+1):
    try:
        leetcode_id = sheet.cell(row=i, column=5).value  
        S_No = sheet.cell(row=i, column=1).value
        RollNo = sheet.cell(row=i, column=2).value
        Name = sheet.cell(row=i, column=4).value
        Class = sheet.cell(row=i, column=3).value
        link = sheet.cell(row=i, column=6).value


        if not all([leetcode_id, S_No, RollNo, Name, Class, link]):
            raise ValueError(f"Missing data in row {i}")

        contest_ranking_query = f"""
        {{
          userContestRanking(username: "{leetcode_id}") {{
            attendedContestsCount
            rating
          }}
        }}
        """
        solved_problems_query = f"""
        {{
          matchedUser(username: "{leetcode_id}") {{
            username
            submitStats: submitStatsGlobal {{
              acSubmissionNum {{
                difficulty
                count
                submissions
              }}
            }}
          }}
        }}
        """

        response_contest_ranking = requests.post(url, json={"query": contest_ranking_query})
        response_solved_problems = requests.post(url, json={"query": solved_problems_query})

        if response_contest_ranking.status_code != 200 or response_solved_problems.status_code != 200:
            raise Exception(f"GraphQL query failed with status {response_contest_ranking.status_code} and {response_solved_problems.status_code}")

        contest_ranking_data = json.loads(response_contest_ranking.text)
        solved_problems_data = json.loads(response_solved_problems.text)

        attended_contests_count = contest_ranking_data['data']['userContestRanking']['attendedContestsCount'] if contest_ranking_data['data']['userContestRanking'] else 'Not Attended'

        rating = contest_ranking_data['data']['userContestRanking']['rating'] if contest_ranking_data['data']['userContestRanking'] else 'No Rating'

        easy_solved = solved_problems_data['data']['matchedUser']['submitStats']['acSubmissionNum'][1]['count'] if solved_problems_data['data']['matchedUser'] else 'Id Wrong'

        medium_solved = solved_problems_data['data']['matchedUser']['submitStats']['acSubmissionNum'][2]['count'] if solved_problems_data['data']['matchedUser'] else 'Id Wrong'

        hard_solved = solved_problems_data['data']['matchedUser']['submitStats']['acSubmissionNum'][3]['count'] if solved_problems_data['data']['matchedUser'] else 'Id Wrong'

    except Exception as e:

        attended_contests_count = 'Error'
        rating = 'Error'
        easy_solved = 'Error'
        medium_solved = 'Error'
        hard_solved = 'Error'
        print(f"Error processing row {i}: {e}")
    new_sheet.append([S_No, RollNo, Name, Class, leetcode_id, link, easy_solved, medium_solved, hard_solved, attended_contests_count, rating])
    time.sleep(1)

# Save the new Excel workbook
new_wb.save('LeetCode_Rishi.xlsx')
files.download('LeetCode_Rishi.xlsx')
